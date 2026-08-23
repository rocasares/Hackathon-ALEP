from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models import ReconciliationReport

HEADER_FONT = Font(bold=True)


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 60)


def _candidates_sheet(wb: Workbook, title: str, candidates: list) -> None:
    ws = wb.create_sheet(title)
    headers = ["Tipo", "Facturas", "Movimientos", "Score", "Decision", "Explicacion"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for c in candidates:
        ws.append(
            [
                c.kind.value if hasattr(c.kind, "value") else c.kind,
                ", ".join(c.invoice_ids),
                ", ".join(c.movement_ids),
                round(c.score, 3),
                c.decision.value if hasattr(c.decision, "value") else c.decision,
                c.explanation,
            ]
        )
    _autosize(ws)


def build_excel(report: ReconciliationReport) -> bytes:
    """Renders a ReconciliationReport as a downloadable .xlsx workbook, one
    sheet per category, plus a summary sheet with the stats block."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Resumen")
    ws_summary.append(["Metrica", "Valor"])
    for cell in ws_summary[1]:
        cell.font = HEADER_FONT
    for key, value in report.stats.items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                ws_summary.append([f"{key}.{sub_key}", sub_value])
        else:
            ws_summary.append([key, value])
    _autosize(ws_summary)

    _candidates_sheet(wb, "Conciliados", report.conciliados)
    _candidates_sheet(wb, "En revision", report.en_revision)

    ws_fsm = wb.create_sheet("Facturas sin movimiento")
    ws_fsm.append(["Factura ID"])
    ws_fsm[1][0].font = HEADER_FONT
    for inv_id in report.facturas_sin_movimiento:
        ws_fsm.append([inv_id])
    _autosize(ws_fsm)

    ws_msc = wb.create_sheet("Movimientos sin comprobante")
    ws_msc.append(["Movimiento ID"])
    ws_msc[1][0].font = HEADER_FONT
    for mov_id in report.movimientos_sin_comprobante:
        ws_msc.append([mov_id])
    _autosize(ws_msc)

    ws_flags = wb.create_sheet("Red flags")
    ws_flags.append(["Tipo", "Facturas", "Razon"])
    for cell in ws_flags[1]:
        cell.font = HEADER_FONT
    for flag in report.red_flags:
        ws_flags.append([flag.get("type", ""), ", ".join(flag.get("invoice_ids", [])), flag.get("reason", "")])
    _autosize(ws_flags)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
